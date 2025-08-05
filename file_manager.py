import os
import json
import hashlib
import sqlite3
from datetime import datetime
from pathlib import Path
import mimetypes
import logging
from typing import List, Dict, Optional, Tuple

# For document processing
try:
    import PyPDF2
    import docx
    import openpyxl
    from bs4 import BeautifulSoup
    import markdown
except ImportError:
    print("Warning: Some document processing libraries are not installed.")
    print("Install with: pip install PyPDF2 python-docx openpyxl beautifulsoup4 markdown")

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DocumentProcessor:
    """Handles different document types and extracts text content"""

    @staticmethod
    def extract_text_from_pdf(file_path: str) -> str:
        """Extract text from PDF files"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text
        except Exception as e:
            logger.error(f"Error processing PDF {file_path}: {e}")
            return ""

    @staticmethod
    def extract_text_from_docx(file_path: str) -> str:
        """Extract text from DOCX files"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error processing DOCX {file_path}: {e}")
            return ""

    @staticmethod
    def extract_text_from_excel(file_path: str) -> str:
        """Extract text from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return text
        except Exception as e:
            logger.error(f"Error processing Excel {file_path}: {e}")
            return ""

    @staticmethod
    def extract_text_from_txt(file_path: str) -> str:
        """Extract text from plain text files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            try:
                with open(file_path, 'r', encoding='cp1251') as file:
                    return file.read()
            except Exception as e:
                logger.error(f"Error processing TXT {file_path}: {e}")
                return ""

    @staticmethod
    def extract_text_from_html(file_path: str) -> str:
        """Extract text from HTML files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                soup = BeautifulSoup(file.read(), 'html.parser')
                return soup.get_text()
        except Exception as e:
            logger.error(f"Error processing HTML {file_path}: {e}")
            return ""

    @staticmethod
    def extract_text_from_md(file_path: str) -> str:
        """Extract text from Markdown files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                md_content = file.read()
                html = markdown.markdown(md_content)
                soup = BeautifulSoup(html, 'html.parser')
                return soup.get_text()
        except Exception as e:
            logger.error(f"Error processing Markdown {file_path}: {e}")
            return ""


class DocumentChunker:
    """Handles chunking of large documents for better processing"""

    def __init__(self, max_chunk_size: int = 4000, overlap_size: int = 200):
        self.max_chunk_size = max_chunk_size
        self.overlap_size = overlap_size

    def chunk_text(self, text: str, document_id: str) -> List[Dict]:
        """Split text into overlapping chunks"""
        chunks = []
        words = text.split()

        if len(words) <= self.max_chunk_size:
            return [{
                'chunk_id': f"{document_id}_chunk_0",
                'content': text,
                'chunk_index': 0,
                'total_chunks': 1
            }]

        total_chunks = (len(words) + self.max_chunk_size - 1) // self.max_chunk_size

        for i in range(0, len(words), self.max_chunk_size - self.overlap_size):
            chunk_words = words[i:i + self.max_chunk_size]
            chunk_text = ' '.join(chunk_words)

            chunks.append({
                'chunk_id': f"{document_id}_chunk_{len(chunks)}",
                'content': chunk_text,
                'chunk_index': len(chunks),
                'total_chunks': total_chunks
            })

            if i + self.max_chunk_size >= len(words):
                break

        return chunks


class FileManager:
    """Enhanced file management system for handling dozens of files"""

    def __init__(self, storage_dir: str = "documents", db_path: str = "file_index.db"):
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(exist_ok=True)
        self.db_path = db_path
        self.processor = DocumentProcessor()
        self.chunker = DocumentChunker()
        self.init_database()

    def init_database(self):
        """Initialize the file index database"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        # Files table
        cursor.execute('''
                       CREATE TABLE IF NOT EXISTS files
                       (
                           id
                           TEXT
                           PRIMARY
                           KEY,
                           filename
                           TEXT
                           NOT
                           NULL,
                           original_name
                           TEXT
                           NOT
                           NULL,
                           file_path
                           TEXT
                           NOT
                           NULL,
                           file_type
                           TEXT
                           NOT
                           NULL,
                           file_size
                           INTEGER
                           NOT
                           NULL,
                           content_hash
                           TEXT
                           NOT
                           NULL,
                           upload_date
                           TIMESTAMP
                           DEFAULT
                           CURRENT_TIMESTAMP,
                           last_modified
                           TIMESTAMP
                           DEFAULT
                           CURRENT_TIMESTAMP,
                           category
                           TEXT,
                           tags
                           TEXT,
                           description
                           TEXT,
                           processed
                           BOOLEAN
                           DEFAULT
                           FALSE,
                           chunk_count
                           INTEGER
                           DEFAULT
                           0
                       )
                       ''')

        # Chunks table for large documents
        cursor.execute('''
                       CREATE TABLE IF NOT EXISTS chunks
                       (
                           id
                           TEXT
                           PRIMARY
                           KEY,
                           file_id
                           TEXT
                           NOT
                           NULL,
                           chunk_index
                           INTEGER
                           NOT
                           NULL,
                           content
                           TEXT
                           NOT
                           NULL,
                           content_preview
                           TEXT,
                           FOREIGN
                           KEY
                       (
                           file_id
                       ) REFERENCES files
                       (
                           id
                       )
                           )
                       ''')

        # Full-text search table
        cursor.execute('''
            CREATE VIRTUAL TABLE IF NOT EXISTS file_search USING fts5(
                file_id,
                filename,
                content,
                category,
                tags
            )
        ''')

        conn.commit()
        conn.close()

    def generate_file_id(self, filename: str) -> str:
        """Generate unique file ID"""
        timestamp = datetime.now().isoformat()
        return hashlib.md5(f"{filename}_{timestamp}".encode()).hexdigest()

    def calculate_file_hash(self, file_path: str) -> str:
        """Calculate file hash for duplicate detection"""
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def detect_file_type(self, file_path: str) -> str:
        """Detect file type based on extension and content"""
        mime_type, _ = mimetypes.guess_type(file_path)
        extension = Path(file_path).suffix.lower()

        type_mapping = {
            '.pdf': 'pdf',
            '.docx': 'docx',
            '.doc': 'doc',
            '.xlsx': 'excel',
            '.xls': 'excel',
            '.txt': 'text',
            '.md': 'markdown',
            '.html': 'html',
            '.htm': 'html',
            '.json': 'json',
            '.xml': 'xml'
        }

        return type_mapping.get(extension, 'unknown')

    def extract_text_content(self, file_path: str, file_type: str) -> str:
        """Extract text content based on file type"""
        extractors = {
            'pdf': self.processor.extract_text_from_pdf,
            'docx': self.processor.extract_text_from_docx,
            'excel': self.processor.extract_text_from_excel,
            'text': self.processor.extract_text_from_txt,
            'html': self.processor.extract_text_from_html,
            'markdown': self.processor.extract_text_from_md
        }

        extractor = extractors.get(file_type, self.processor.extract_text_from_txt)
        return extractor(file_path)

    def upload_file(self, file_path: str, category: str = None, tags: List[str] = None,
                    description: str = None) -> Dict:
        """Upload and process a file"""
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            # Generate file info
            file_id = self.generate_file_id(file_path.name)
            file_type = self.detect_file_type(str(file_path))
            file_size = file_path.stat().st_size
            content_hash = self.calculate_file_hash(str(file_path))

            # Copy file to storage
            storage_path = self.storage_dir / f"{file_id}_{file_path.name}"
            storage_path.write_bytes(file_path.read_bytes())

            # Extract text content
            text_content = self.extract_text_content(str(storage_path), file_type)

            # Chunk large documents
            chunks = self.chunker.chunk_text(text_content, file_id)

            # Store in database
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            # Insert file record
            cursor.execute('''
                           INSERT INTO files (id, filename, original_name, file_path, file_type,
                                              file_size, content_hash, category, tags, description,
                                              processed, chunk_count)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                           ''', (
                               file_id, file_path.name, str(file_path), str(storage_path),
                               file_type, file_size, content_hash, category,
                               json.dumps(tags or []), description, True, len(chunks)
                           ))

            # Insert chunks
            for chunk in chunks:
                cursor.execute('''
                               INSERT INTO chunks (id, file_id, chunk_index, content, content_preview)
                               VALUES (?, ?, ?, ?, ?)
                               ''', (
                                   chunk['chunk_id'], file_id, chunk['chunk_index'],
                                   chunk['content'], chunk['content'][:200] + "..."
                               ))

                # Add to search index - only for non-problematic content
                try:
                    cursor.execute('''
                                   INSERT INTO file_search (file_id, filename, content, category, tags)
                                   VALUES (?, ?, ?, ?, ?)
                                   ''', (
                                       file_id, file_path.name, chunk['content'],
                                       category or '', json.dumps(tags or [])
                                   ))
                except Exception as search_error:
                    logger.warning(f"FTS5 index error for chunk {chunk['chunk_id']}: {search_error}")
                    # Continue without FTS5 indexing for this chunk

            conn.commit()
            conn.close()

            logger.info(f"Successfully uploaded and processed: {file_path.name}")
            return {
                'file_id': file_id,
                'filename': file_path.name,
                'file_type': file_type,
                'chunks': len(chunks),
                'success': True
            }

        except Exception as e:
            logger.error(f"Error uploading file {file_path}: {e}")
            return {'success': False, 'error': str(e)}

    def clean_search_query(self, query: str) -> str:
        """Clean search query to avoid FTS5 syntax errors"""
        import re

        # Remove special characters that cause FTS5 issues
        query = query.replace('"', '')
        query = query.replace("'", '')
        query = query.replace('?', '')
        query = query.replace('(', '')
        query = query.replace(')', '')
        query = query.replace('[', '')
        query = query.replace(']', '')
        query = query.replace('{', '')
        query = query.replace('}', '')
        query = query.replace('*', '')
        query = query.replace('+', '')
        query = query.replace('-', ' ')

        # Split into words and rejoin
        words = query.split()
        cleaned_words = []

        for word in words:
            if len(word) >= 2:  # Only include words with 2+ characters
                cleaned_words.append(word)

        return ' '.join(cleaned_words) if cleaned_words else query

    def fallback_search(self, query: str, category: str = None, file_type: str = None) -> List[Dict]:
        """Fallback search using simple LIKE queries"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            search_query = """
                           SELECT DISTINCT f.id, \
                                           f.filename, \
                                           f.file_type, \
                                           f.category, \
                                           f.description,
                                           f.chunk_count, \
                                           SUBSTR(c.content, 1, 300) as snippet
                           FROM files f
                                    JOIN chunks c ON f.id = c.file_id
                           WHERE (c.content LIKE ? OR f.filename LIKE ? OR f.description LIKE ?) \
                           """
            params = [f"%{query}%", f"%{query}%", f"%{query}%"]

            if category:
                search_query += " AND f.category = ?"
                params.append(category)

            if file_type:
                search_query += " AND f.file_type = ?"
                params.append(file_type)

            search_query += " ORDER BY f.upload_date DESC LIMIT 20"

            cursor.execute(search_query, params)
            results = cursor.fetchall()

            search_results = []
            for row in results:
                search_results.append({
                    'file_id': row[0],
                    'filename': row[1],
                    'file_type': row[2],
                    'category': row[3],
                    'description': row[4],
                    'chunk_count': row[5],
                    'snippet': row[6] if row[6] else ""
                })

            conn.close()
            return search_results

        except Exception as e:
            logger.error(f"Fallback search error: {e}")
            conn.close()
            return []

    def search_files(self, query: str, category: str = None, file_type: str = None) -> List[Dict]:
        """Search through all files and their content - FIXED VERSION"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            # Clean the query to avoid FTS5 syntax errors
            cleaned_query = self.clean_search_query(query)

            if not cleaned_query.strip():
                return self.fallback_search(query, category, file_type)

            # Check if query contains Azerbaijani characters or problematic symbols
            has_azerbaijani = any(char in query for char in ['ə', 'ı', 'ö', 'ü', 'ğ', 'ş', 'ç'])
            has_special_chars = any(char in query for char in ['"', "'", '?', '(', ')', '[', ']'])

            if has_azerbaijani or has_special_chars:
                # Use LIKE search for Azerbaijani or special characters
                return self.fallback_search(query, category, file_type)

            # Try FTS5 search for simple queries
            try:
                search_query = """
                               SELECT DISTINCT f.id, \
                                               f.filename, \
                                               f.file_type, \
                                               f.category, \
                                               f.description,
                                               f.chunk_count, \
                                               snippet(file_search, 2, '<mark>', '</mark>', '...', 32) as snippet
                               FROM files f
                                        JOIN file_search fs ON f.id = fs.file_id
                               WHERE file_search MATCH ? \
                               """
                params = [cleaned_query]

                # Add category filter if specified
                if category:
                    search_query += " AND f.category = ?"
                    params.append(category)

                # Add file type filter if specified
                if file_type:
                    search_query += " AND f.file_type = ?"
                    params.append(file_type)

                search_query += " ORDER BY f.upload_date DESC LIMIT 20"

                cursor.execute(search_query, params)
                results = cursor.fetchall()

                search_results = []
                for row in results:
                    search_results.append({
                        'file_id': row[0],
                        'filename': row[1],
                        'file_type': row[2],
                        'category': row[3],
                        'description': row[4],
                        'chunk_count': row[5],
                        'snippet': row[6] if row[6] else ""
                    })

                conn.close()
                return search_results

            except Exception as fts_error:
                logger.warning(f"FTS5 search failed: {fts_error}, falling back to LIKE search")
                conn.close()
                return self.fallback_search(query, category, file_type)

        except Exception as e:
            logger.error(f"Search error: {e}")
            conn.close()
            return self.fallback_search(query, category, file_type)

    def get_file_content(self, file_id: str, chunk_index: int = None) -> Dict:
        """Get file content, optionally specific chunk"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        if chunk_index is not None:
            cursor.execute('''
                           SELECT content
                           FROM chunks
                           WHERE file_id = ?
                             AND chunk_index = ?
                           ''', (file_id, chunk_index))
            result = cursor.fetchone()
            content = result[0] if result else ""
        else:
            cursor.execute('''
                           SELECT content
                           FROM chunks
                           WHERE file_id = ?
                           ORDER BY chunk_index
                           ''', (file_id,))
            chunks = cursor.fetchall()
            content = "\n\n".join([chunk[0] for chunk in chunks])

        # Get file info
        cursor.execute('''
                       SELECT filename, file_type, category, description, chunk_count
                       FROM files
                       WHERE id = ?
                       ''', (file_id,))
        file_info = cursor.fetchone()

        conn.close()

        if file_info:
            return {
                'content': content,
                'filename': file_info[0],
                'file_type': file_info[1],
                'category': file_info[2],
                'description': file_info[3],
                'chunk_count': file_info[4]
            }
        return {'error': 'File not found'}

    def list_files(self, category: str = None) -> List[Dict]:
        """List all uploaded files"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        if category:
            cursor.execute('''
                           SELECT id,
                                  filename,
                                  file_type,
                                  file_size,
                                  category,
                                  description,
                                  upload_date,
                                  chunk_count
                           FROM files
                           WHERE category = ?
                           ORDER BY upload_date DESC
                           ''', (category,))
        else:
            cursor.execute('''
                           SELECT id,
                                  filename,
                                  file_type,
                                  file_size,
                                  category,
                                  description,
                                  upload_date,
                                  chunk_count
                           FROM files
                           ORDER BY upload_date DESC
                           ''')

        files = []
        for row in cursor.fetchall():
            files.append({
                'file_id': row[0],
                'filename': row[1],
                'file_type': row[2],
                'file_size': row[3],
                'category': row[4],
                'description': row[5],
                'upload_date': row[6],
                'chunk_count': row[7]
            })

        conn.close()
        return files

    def bulk_upload(self, directory_path: str, category: str = None) -> Dict:
        """Upload all files from a directory"""
        directory = Path(directory_path)
        if not directory.exists():
            return {'error': 'Directory not found'}

        results = {'successful': [], 'failed': []}
        supported_extensions = {'.pdf', '.docx', '.xlsx', '.txt', '.md', '.html'}

        for file_path in directory.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                result = self.upload_file(str(file_path), category=category)
                if result.get('success'):
                    results['successful'].append(result)
                else:
                    results['failed'].append({'file': str(file_path), 'error': result.get('error')})

        return {
            'total_processed': len(results['successful']) + len(results['failed']),
            'successful': len(results['successful']),
            'failed': len(results['failed']),
            'details': results
        }